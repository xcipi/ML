# -*- coding: utf-8 -*-
"""
Created on Mon Jan 17 13:20:54 2017
@author: skipi

tools and functions for internal ML tests

"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import (
    get_column_letter,
    column_index_from_string,
)

def printCells(ws,lu,rd):
    """
    tuto funkciu nakoniec nepouzivam
    """
    try:
        for rowOfCellObjects in ws[lu:rd]:
            for cellObj in rowOfCellObjects:
                print(cellObj.coordinate, cellObj.value)
            print('--- END OF ROW ---')
    except Exception as e:
        print ('ERROR!', e)

def findFilteredMetaRow(ws,type_id):
    '''
    funkcia zisti v metaData sheete, ktory riadok pouzit na zaklade type_id
    '''
    filteredRow = 0
    for row in range (1,ws.max_row+1):
        if ws['A' + str(row)].value == type_id:
            filteredRow = row
    return(filteredRow)

def findFilteredSubject(ws,row):
    '''
    funkcia vrati hodnotu filtrovaneho subjectu
    '''
    return(ws['B' + str(row)].value)

def fillMetaDataDict(ws):
    '''
    vytvori a naplni dictionary z meta dat, pomocou ktorej potom rozhodujem, ktore hodnoty vypisovat a ktore nie (filter), key je subject_filter
    '''
    fillMetaData = {}
    for row in range (3,ws.max_row+1):
        fillMetaData.setdefault(ws['B' + str(row)].value, {})
        for col in range (1,ws.max_column+1):
            fillMetaData[ws['B' + str(row)].value].update({str(ws[get_column_letter(col)+str(2)].value):str(ws[get_column_letter(col)+str(row)].value)})
    return(fillMetaData)

def fillFilteredDataDict(ws, metaData, subject):
    '''
    vytvori disctionary s realnymi datami z Data sheetu (ws), filtrovanymi podla MetaData (1|0). key je v tomto pripade ID (jedina jedinecna hodnota) 
    '''
    filteredData = {}
    for row in range (3,ws.max_row+1):
        if ws['A' + str(row)].value == subject:
            filteredData.setdefault(ws['D' + str(row)].value, {})
            for col in range (5,ws.max_column+1):
                if metaData[subject][ws[get_column_letter(col) + str(2)].value] == str(1):
                    filteredData[ws['D' + str(row)].value].update({str(ws[get_column_letter(col)+str(2)].value):str(ws[get_column_letter(col)+str(row)].value)})
    return(filteredData)


def getDataByFilter(type_id):
    '''
    funkcia, ktoru som pouzil miesto main(), lebo som nevedel, ci to chcete niekde pouzit ako modul alebo to pouzijete priamo. 
    '''
    #XlsFile = 'M:/IOBox/BigML_DataSourceBase.xlsx'
    XlsFile = '\\\\lynxhafile\PBR_DATA/PBR DATA/kBase/ML/DataSets/VPN_Events_sample2066_02.xlsx'
    WsMetaName = 'MetaData'
    WsDataName = 'Data'

    metaData = {}
    filteredData = {}
    
    wb = load_workbook(XlsFile)
    wsm = wb.get_sheet_by_name(WsMetaName)
    wsd = wb.get_sheet_by_name(WsDataName)

    wsm_mr = wsm.max_row
    wsd_mr = wsd.max_row
    wsm_mc = wsm.max_column
    wsd_mc = wsd.max_column

    subject = findFilteredSubject(wsm,findFilteredMetaRow(wsm,type_id))
    metaData = fillMetaDataDict(wsm)

    filteredData = fillFilteredDataDict(wsd, metaData, subject)
    print (filteredData)

    '''
    testovanie volania funkcie s parametrom type_id
    '''

getDataByFilter('X1')
#getDataByFilter('AC')