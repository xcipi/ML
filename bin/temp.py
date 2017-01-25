# -*- coding: utf-8 -*-
"""
Created on Mon Jan 17 13:20:54 2017
@author: skipi

tools and functions for internal ML tests

"""

from openpyxl import Workbook
from openpyxl import load_workbook
import re

#WSregex = r"^(Prod|Test|MGMT).*$"
WSregex = r"^(Prod)|(Test)|(MGMT).*$"
XlsFile = './SKI-RPI-IPv4 segmentacia.xlsx'
itermatch = 0
iterother = 0

try:
    wb2 = load_workbook(XlsFile, read_only=True)
    for i in wb2.get_sheet_names():
        matches = re.finditer(WSregex, i)
        for matchNum, match in enumerate(matches):
            print ( 'Match: ', matchNum)
            matchNum = matchNum + 1
            print(match.group())
            for groupNum in range(0, len(match.groups())):
                iterother=iterother+1
                groupNum = groupNum + 1
                matchGroup = match.group(groupNum)
                if matchGroup: 
                    print ('ITERother#',iterother,' ',match.group(groupNum))
                    print ('END')                
                
                
except Exception as e:
    print ('ERROR!', e)
'''
iter=0
wb2 = load_workbook(XlsFile, read_only=True)
for i in wb2.get_sheet_names():
    iter = iter+1
    matches = re.finditer(WSregex, i)
    for matchNum, match in enumerate(matches):
        matchNum = matchNum + 1
        print ("ITER ",iter, "Match {matchNum} was found at {start}-{end}: {match}".format(matchNum = matchNum, start = match.start(), end = match.end(), match = match.group()))
        for groupNum in range(0, len(match.groups())):
            groupNum = groupNum + 1
            print ("Group {groupNum} found at {start}-{end}: {group}".format(groupNum = groupNum, start = match.start(groupNum), end = match.end(groupNum), group = match.group(groupNum)))

''' 
'''
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save("pandas_openpyxl.xlsx")
'''
