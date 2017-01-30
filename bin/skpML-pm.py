# -*- coding: utf-8 -*-
"""
Created on Mon Jan 17 13:20:54 2017
@author: skipi
version 2017012402

tools and functions for internal ML tests

"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import (
    get_column_letter,
    column_index_from_string,
)

def printCells(ws,lu,rd):
    """
    tuto funkciu nakoniec nepouzivam
    """
    try:
        for rowOfCellObjects in ws[lu:rd]:
            for cellObj in rowOfCellObjects:
                print(cellObj.coordinate, cellObj.value)
            print('--- END OF ROW ---')
    except Exception as e:
        print ('ERROR!', e)

def findFilteredMetaRow(ws,type_id):
    '''
    funkcia zisti v metaData sheete, ktory riadok pouzit na zaklade type_id
    '''
    filteredRow = 0
    for row in range (1,ws.max_row+1):
        if ws['A' + str(row)].value == type_id:
            filteredRow = row
    return(filteredRow)

def findFilteredSubject(ws,row):
    '''
    funkcia vrati hodnotu filtrovaneho subjectu
    '''
    return(ws['B' + str(row)].value)

def fillMetaDataDict(ws):
    '''
    vytvori a naplni dictionary z meta dat, pomocou ktorej potom rozhodujem, ktore hodnoty vypisovat a ktore nie (filter), key je subject_filter
    '''
    fillMetaData = {}
    for row in range (3,ws.max_row+1):
        fillMetaData.setdefault(ws['B' + str(row)].value, {})
        for col in range (1,ws.max_column+1):
            fillMetaData[ws['B' + str(row)].value].update({str(ws[get_column_letter(col)+str(2)].value):str(ws[get_column_letter(col)+str(row)].value)})
    return(fillMetaData)

def fillFilteredDataDict(ws, metaData, subject):
    '''
    vytvori disctionary s realnymi datami z Data sheetu (ws), filtrovanymi podla MetaData (1|0). key je v tomto pripade ID (jedina jedinecna hodnota) 
    '''
    filteredData = {}
#    print ('# subject: ', subject)
    for row in range (3,ws.max_row+1):
#        print ('# row: ',row,'value in WS: ',ws['A' + str(row)].value,' ###')
        if ws['A' + str(row)].value == subject:
            filteredData.setdefault(ws['D' + str(row)].value, {})
#            print ('## row: ', row, ' matching ...')
            for col in range (5,ws.max_column+1):
#                print ('## col: ',col,' value ',ws[get_column_letter(col)+str(row)].value,' ###')
#                print ('subject: ',subject,' col letter: ',get_column_letter(col), ' matching row: ', str(1))
#                print (metaData[subject][ws[get_column_letter(col) + str(1)].value])
                if metaData[subject][ws[get_column_letter(col) + str(2)].value] == str(1):
#                    print ('### ', metaData[subject][ws[get_column_letter(col) + str(1)]])
                    filteredData[ws['D' + str(row)].value].update({str(ws[get_column_letter(col)+str(2)].value):str(ws[get_column_letter(col)+str(row)].value)})
    return(filteredData)


def getDataByFilter(type_id):
    '''
    funkcia, ktoru som pouzil miesto main(), lebo som nevedel, ci to chcete niekde pouzit ako modul alebo to pouzijete priamo. 
    '''
    XlsFile = '\\\\lynxhafile\PBR_DATA/PBR DATA/kBase/ML/DataSets/VPN_Events_TestData.xlsx'
    WsMetaName = 'MetaData'
    WsDataName = 'Data'

    metaData = {}
    filteredData = {}
    
    print ('### loading workbook ...')
    wb = load_workbook(XlsFile, data_only=True)
    print ('### loading worksheet ',WsMetaName,' ...')
    wsm = wb.get_sheet_by_name(WsMetaName)
    print ('### loading worksheet ',WsDataName,' ...')
    wsd = wb.get_sheet_by_name(WsDataName)

    wsm_mr = wsm.max_row
    wsd_mr = wsd.max_row
    wsm_mc = wsm.max_column
    wsd_mc = wsd.max_column
    
    print ('### Searching for subject for: ',type_id, ' ...')
    subject = findFilteredSubject(wsm,findFilteredMetaRow(wsm,type_id))
    print ('#### Subject is ', subject)
    print ('### Filling metadata dictionary ...')
    metaData = fillMetaDataDict(wsm)

    filteredData = fillFilteredDataDict(wsd, metaData, subject)
#    print (filteredData)
    return(filteredData)


from bigml.api import BigML
from bigml.anomaly import Anomaly
from bigml.cluster import Cluster
import os
import requests

proxy_server='10.61.3.100:8080'
proxy_username = 'LYNX_NT\peterm'
#proxy_pwd = ''

proxy_string = proxy_username + ':' + proxy_pwd + '@' + proxy_server

proxies = {
    'http': 'http://'+proxy_string,
    'https': 'https://'+proxy_string,
    'ftp': 'ftp://'+proxy_string
}

os.environ['HTTPS_PROXY'] = 'https://'+proxy_string

r = requests.get('https://bigml.io') #, proxies = proxies)
if r.status_code == requests.codes.ok:
    print(r.headers['content-type'])

print ('### Connecting to the BigML ...')

api = BigML('peterm-lynx_sk', '5719b317de624a074eb3650e3b527133bedb29ee', 
            dev_mode=True, domain="bigml.io")

Type_ID = 'X3_155'
VerNo = '-v5'

filteredData = getDataByFilter(Type_ID)

filteredDataList = []
for elem in filteredData:
    filteredDataList.append(filteredData[elem])

print ('### Creating data Source ...')
source = api.create_source(filteredDataList, {'name': Type_ID + VerNo})
#api.ok(source)

print ('### Creating Dataset ...')
dataset = api.create_dataset(source, {"name": Type_ID + VerNo})
#api.ok(dataset)

print ('### Creating Anomaly detector ...')
anomaly = api.create_anomaly(dataset, {"name": Type_ID + VerNo})


