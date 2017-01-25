from openpyxl import Workbook
from openpyxl import load_workbook
import re






from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# add a simple formula
ws["A1"] = "=SUM(1, 1)"
print ('### ', ws["A1"], ' # ', ws["A1"].value)
wb.save("formula.xlsx")

